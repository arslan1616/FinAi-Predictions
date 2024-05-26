import openpyxl
from pathlib import Path
from pmdarima import auto_arima
import numpy as np
from sklearn.metrics import mean_absolute_percentage_error
from prophet import Prophet
import pandas as pd
import tensorflow as tf
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import LSTM, Dense, Input
import matplotlib.pyplot as plt
import os
import re

def read_equity_data_from_excel(folder_path, keywords):
    pathlist = Path(folder_path).rglob('*.xlsx')
    data_dict = {}
    for path in pathlist:
        path_in_str = str(path)
        wb = openpyxl.load_workbook(path_in_str, data_only=True)
        sheet = wb.active
        data = {}
        for row in sheet.iter_rows(values_only=True):
            if row and isinstance(row[0], str):
                for keyword in keywords:
                    if row[0].strip() == keyword.strip():
                        values = list(row[1:])
                        values.reverse()
                        values = [value for value in values if value is not None and not np.isnan(value) and value != 0]
                        if values:
                            if keyword not in data:
                                data[keyword] = []
                            data[keyword].append(values)
        data_dict[path_in_str] = data
    return data_dict

def convert_to_float(data_dict):
    for path, data in data_dict.items():
        for key, value_lists in data.items():
            data_dict[path][key] = [[float(value) for value in values] for values in value_lists]
    return data_dict

def lstm_forecast(values):
    if len(values) < 2:
        return [np.nan] * len(values)

    values = np.array(values)
    values = values.reshape((values.shape[0], 1, 1))
    model = Sequential()
    model.add(Input(shape=(1, 1)))
    model.add(LSTM(50, activation='relu'))
    model.add(Dense(1))
    model.compile(optimizer='adam', loss='mse')
    try:
        model.fit(values[:-1], values[1:], epochs=300, verbose=0)
    except tf.errors.OutOfRangeError:
        return [np.nan] * len(values)

    forecasts = model.predict(values, verbose=0)
    return forecasts.flatten()

def prophet_forecast(values):
    if len(values) < 2:
        return [np.nan] * len(values)

    df = pd.DataFrame({'ds': pd.date_range(start='2020-01-01', periods=len(values), freq='Q'), 'y': values})
    model = Prophet()
    model.fit(df)
    future = model.make_future_dataframe(periods=len(values), freq='Q')
    forecast = model.predict(future)
    return forecast['yhat'].values[:len(values)]

def predict_next_value(data_dict):
    predictions = {}
    for path, data in data_dict.items():
        predictions[path] = {}
        for key, value_lists in data.items():
            predictions[path][key] = []
            for values in value_lists:
                if len(values) >= 4:
                    try:
                        forecasts = {
                            'arima': [],
                            'lstm': [],
                            'prophet': []
                        }

                        # ARIMA forecast
                        model_arima = auto_arima(values, seasonal=False, stepwise=True, suppress_warnings=True,
                                                 error_action="ignore", trace=False)
                        forecast_arima = model_arima.predict(n_periods=len(values))
                        mape_arima = mean_absolute_percentage_error(values, forecast_arima)

                        best_forecast = forecast_arima[-1]
                        best_mape = mape_arima

                        # LSTM forecast
                        forecast_lstm = lstm_forecast(values)
                        if not np.isnan(forecast_lstm).any():
                            mape_lstm = mean_absolute_percentage_error(values, forecast_lstm)
                            if mape_lstm < best_mape:
                                best_mape = mape_lstm
                                best_forecast = forecast_lstm[-1]

                        # Prophet forecast
                        forecast_prophet = prophet_forecast(values)
                        if not np.isnan(forecast_prophet).any():
                            mape_prophet = mean_absolute_percentage_error(values, forecast_prophet)
                            if mape_prophet < best_mape:
                                best_mape = mape_prophet
                                best_forecast = forecast_prophet[-1]

                        last_value = values[-1]
                        percentage_change = ((best_forecast - last_value) / last_value) * 100 if last_value != 0 else 0
                        predictions[path][key].append((best_forecast, percentage_change))

                        # Save the forecast visualization
                        save_forecast_visualization(values, forecast_arima, forecast_lstm, forecast_prophet, key, path)

                    except Exception as e:
                        predictions[path][key].append((np.nan, np.nan))
                        print(f"Error processing {key} in {path}: {e}")
                else:
                    print(f"Skipping {key} in {path} due to insufficient data length ({len(values)} values).")
                    predictions[path][key].append((np.nan, np.nan))
    return predictions

def sanitize_filename(filename):
    return re.sub(r'[\\/*?:"<>|]', "_", filename)

def save_forecast_visualization(values, forecast_arima, forecast_lstm, forecast_prophet, key, path):
    folder_name = os.path.basename(path).replace('.xlsx', '')
    save_path = os.path.join('forecast_plots', folder_name)
    os.makedirs(save_path, exist_ok=True)

    sanitized_key = sanitize_filename(key)

    plt.figure(figsize=(12, 6))
    plt.plot(range(len(values)), values, label='Gerçek Değerler')
    plt.plot(range(len(forecast_arima)), forecast_arima, color='r', linestyle='--', label='ARIMA Tahmini')
    plt.plot(range(len(forecast_lstm)), forecast_lstm, color='g', linestyle='--', label='LSTM Tahmini')
    plt.plot(range(len(forecast_prophet)), forecast_prophet, color='b', linestyle='--', label='Prophet Tahmini')
    plt.xlabel('Zaman')
    plt.ylabel('Değer')
    plt.title(f'{key} - {folder_name}')
    plt.legend()
    plt.savefig(os.path.join(save_path, f'{sanitized_key}.png'))
    plt.close()

def update_excel_with_predictions(folder_path, data_dict, predictions):
    pathlist = Path(folder_path).rglob('*.xlsx')
    for path in pathlist:
        path_in_str = str(path)
        wb = openpyxl.load_workbook(path_in_str)
        sheet = wb.active

        for key in data_dict[path_in_str]:

            headers = [cell.value for cell in sheet[1]]
            if "Tahmin" not in headers:
                prediction_col = sheet.max_column + 1
                sheet.cell(row=1, column=prediction_col, value="Tahmin")
            else:
                prediction_col = headers.index("Tahmin") + 1

            if "Tahmini Değişim" not in headers:
                change_col = sheet.max_column + 1
                sheet.cell(row=1, column=change_col, value="Tahmini Değişim")
            else:
                change_col = headers.index("Tahmini Değişim") + 1

            for row_index, row in enumerate(sheet.iter_rows(min_col=1, max_col=1, values_only=True), start=1):
                if row:
                    cell_value = str(row[0])
                    if cell_value == key:
                        for i, (predicted_value, percentage_change) in enumerate(predictions[path_in_str][key]):
                            sheet.cell(row=row_index + i, column=prediction_col, value=predicted_value)
                            sheet.cell(row=row_index + i, column=change_col, value=percentage_change)

        wb.save(path_in_str)

folder_path = 'D:/dolarbilanco'
keywords = ["Özkaynaklar", "Net Faaliyet Kar/Zararı"," Özsermaye Toplamı", "  F-Dönem Net Karı ",
            "XVI. ÖZKAYNAKLAR", "XVII. SÜRDÜRÜLEN FAALİYETLER DÖNEM NET K/Z (XV±XVI)", "ÖZKAYNAK"]
data = read_equity_data_from_excel(folder_path, keywords)
data = convert_to_float(data)
predictions = predict_next_value(data)
update_excel_with_predictions(folder_path, data, predictions)
